import csv
import inspect
import logging
from openpyxl import Workbook, load_workbook


class LogtheAutomation():
    def logcatch(loglevel=logging.DEBUG):
        logger_name = inspect.stack()[1][3]
        # create Logger
        logger = logging.getLogger(logger_name)

        if not logger.handlers:
            # setlevel
            logger.setLevel(logging.DEBUG)

            # create a handler for  consle or file handler
            # consolehandler = logging.StreamHandler()
            filehandler = logging.FileHandler("automation.log")

            # create fromaterr- how you want your logs to be formatted that u need
            # format1 = logging.Formatter('%(asctime)s %(message)s %(name)s')
            format2 = logging.Formatter('%(asctime)s %(message)s', datefmt='%m/%d/%Y %I:%M:%S %p')

            # add your formatter to file or console handler
            # consolehandler.setFormatter(format1)
            filehandler.setFormatter(format2)

            # add those handler to logger
            # logger.addHandler(consolehandler)
            logger.addHandler(filehandler)
        return logger


    def xlsxfilereader(excelname,sheetname):
        datalist =[]
        excelfile = load_workbook(filename=excelname)
        sheet = excelfile[sheetname]
        roww = sheet.max_row
        # roww = sheet.cell(1,3)
        coll = sheet.max_column

        for i in range(2,roww+1): #skiping the header
            store =[]
            for j in range(1,coll+1):
                store.append(sheet.cell(row=i, column=j).value)
            datalist.append(store)
        return datalist

    def csvfilereader(filename):
        datalist=[]
        csvfile = open(filename,"r")   #opencsvfile
        reader = csv.reader(csvfile) #reading the csv file
        next(reader)  #skiping the header
        for rows in reader:
            datalist.append(rows)
        return datalist




